from CDMHandler import CDMHandler
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def styled_cells(ws, data):
    #Loop through the data items
    i=1
    for col in data:
      c = ws.cell(row = 1, column=i, value=col)
      c.font = Font(bold=True)
      i+=1
      yield c

def autofit_cells(ws):
  for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length+2

wb = Workbook()
ws1 = wb.active
if __name__ == '__main__' :
  import sys
  if len(sys.argv) <= 2:
    print("USAGE:   ",sys.argv[0],"<Powerdesigner pdm path/file> <output Excel file>")
    print("EXAMPLE: ",sys.argv[0],"data/Consol.pdm Consol_datadict.xls")
    sys.exit(1)
  else:
    filename = sys.argv[1]
    dest_filename = sys.argv[2]
  ph = CDMHandler.parse(filename)

  ws1.title = "Packages"
  ws1.append(styled_cells(ws1, ['Name','Code','Creator']))

  ws2 = wb.create_sheet("Entities")
  ws2.append(styled_cells(ws2, ["Package Name", "Entity Name", "Code", "Creator"]))
  for pkg in CDMHandler.getPkgNodes(ph):
    pkg_attrs = CDMHandler.getPkgAttrs(pkg)
    ws1.append(list([pkg_attrs["Name"],pkg_attrs["Code"],pkg_attrs["Creator"]]))
#    print("P:", pkg_attrs["Name"],pkg_attrs["Code"],datetime.fromtimestamp(int(pkg_attrs["CreationDate"])), pkg_attrs["Creator"])
    for ent in CDMHandler.getEntNodesInPkg(pkg) :
      ent_attrs = CDMHandler.getEntAttrs(ent)
      ws2.append([pkg_attrs["Name"], ent_attrs["Name"],ent_attrs["Code"],ent_attrs["Creator"]])
#      print(" E:", ent_attrs["Name"],ent_attrs["Code"],datetime.fromtimestamp(int(ent_attrs["CreationDate"])), ent_attrs["Creator"])
  autofit_cells(ws1)
  autofit_cells(ws2)
  wb.save(filename=dest_filename)