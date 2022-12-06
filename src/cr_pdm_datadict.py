from PDMHandler import PDMHandler
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def styled_cells(ws, data):
    # Loop through the data items
    i = 1
    for col in data:
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(bold=True)
        i += 1
        yield c


def autofit_cells(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


wb = Workbook()
ws1 = wb.active
if __name__ == '__main__':
    import sys

    if len(sys.argv) <= 2:
        print("USAGE:   ", sys.argv[0], "<Powerdesigner pdm path/file> <output Excel file>")
        print("EXAMPLE: ", sys.argv[0], "data/Consol.pdm Consol_datadict.xls")
        sys.exit(1)
    else:
        filename = sys.argv[1]
        dest_filename = sys.argv[2]
    ph = PDMHandler.parse(filename)
    ws1.title = "Packages"
    ws1.append(styled_cells(ws1, ['Name', 'Code', 'Creator']))

    ws2 = wb.create_sheet("Tables")
    ws2.append(styled_cells(ws2, ["Package Name", "Table Name", "Code", "Creator"]))

    ws3 = wb.create_sheet("Table Cols")
    ws3.append(
        styled_cells(ws3, ["Package Name", "Table Name", "Column Name", "Code", "DataType", "Length", "Mandatory?"]))

    ws4 = wb.create_sheet("Indexes")
    ws4.append(styled_cells(ws4, ["Package Name", "Table Name", "Index Name", "Code", "Unique?"]))

    ws5 = wb.create_sheet("Index Cols")
    ws5.append(styled_cells(ws5, ["Package Name", "Table Name", "Index Name", "RefColCode"]))

    for pkg in PDMHandler.getPkgNodes(ph):
        pkg_attrs = PDMHandler.getPkgAttrs(pkg)
        ws1.append(list([pkg_attrs["Name"], pkg_attrs["Code"], pkg_attrs["Creator"]]))
        #   print("P:", pkg_attrs[ws = wb.create_sheet()"Name"],pkg_attrs["Code"],pkg_attrs["Creator"])

        for tbl in PDMHandler.getTblNodesInPkg(pkg):
            tbl_attrs = PDMHandler.getTblAttrs(tbl)
            ws2.append([pkg_attrs["Name"], tbl_attrs["Name"], tbl_attrs["Code"], tbl_attrs["Creator"]])
            # print(" T:", tbl_attrs["Name"],tbl_attrs["Code"],tbl_attrs["Creator"])
            # print("  T-PATH:",PDMHandler.getNodePath(tbl))
            for col in PDMHandler.getColNodesInTbl(tbl):
                col_attrs = PDMHandler.getColAttrs(col)
                ws3.append(
                    [pkg_attrs["Name"], tbl_attrs["Name"], col_attrs["Name"], col_attrs["Code"], col_attrs["DataType"],
                     col_attrs["Length"], col_attrs["Column.Mandatory"]])
                # print("  C:", col_attrs["Name"],col_attrs["Code"],col_attrs["DataType"],col_attrs["Length"],col_attrs["Column.Mandatory"])
            for idx in PDMHandler.getIdxNodesInTbl(tbl):
                idx_attrs = PDMHandler.getIdxAttrs(idx)
                ws4.append(
                    [pkg_attrs["Name"], tbl_attrs["Name"], idx_attrs["Name"], idx_attrs["Code"], idx_attrs["Unique"]])
                # print("  I:", idx_attrs["Name"],idx_attrs["Code"],idx_attrs["Unique"])
                for idxcol in PDMHandler.getIdxColNodesInIdx(idx):
                    idxcol_attrs = PDMHandler.getIdxColAttrs(idxcol)
                    ws5.append([pkg_attrs["Name"], tbl_attrs["Name"], idx_attrs["Name"], idxcol_attrs["RefColCode"]])
                    # print("   IC:", idxcol_attrs["RefColCode"])
            autofit_cells(ws1)
            autofit_cells(ws2)
            autofit_cells(ws3)
            autofit_cells(ws4)
            autofit_cells(ws5)
            wb.save(filename=dest_filename)
